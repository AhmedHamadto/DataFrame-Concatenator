import pandas as pd
import openpyxl
import glob
import os
import xlsxwriter
from openpyxl import load_workbook
from tkinter import Tk, filedialog
from pathlib import Path



def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs)
        return

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

root = Tk()
root.withdraw()

root.attributes('-topmost', True)

source_dir = filedialog.askdirectory()

file_names = glob.glob(os.path.join(source_dir, '*.xlsx'))

service_name = os.path.basename(os.path.dirname(source_dir))

year = os.path.basename(source_dir)

target_file = service_name + " " + year + ".xlsx"

workbook = xlsxwriter.Workbook(target_file)
worksheet = workbook.add_worksheet()

workbook.close()



srcfile = openpyxl.load_workbook(target_file, read_only=False,
                                         keep_vba=False)

sheetname = srcfile['Sheet1']

r = 1
#
# sheetname["A1"] = "Well"
# sheetname["B1"] = "Well Type"
# sheetname["C1"] = "Rig Activity"
# sheetname["D1"] = "Month"

srcfile.save(target_file)

for file in file_names:


    df=pd.read_excel(file)
    # print (df)
    # df.drop(["Number","Number of days","Number of callouts"],
    #         axis = 1, inplace = True)

    append_df_to_excel(target_file,df,header = False, index = False, startrow = r, startcol = 0)

    index = df.index
    Num_of_Rigs = len(index)
    r += Num_of_Rigs
    print (r)